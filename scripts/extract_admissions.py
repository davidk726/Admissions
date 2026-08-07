#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Extract normalized admission-result records from university files."""

from __future__ import annotations

import json
import re
import sys
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

try:
    import pymupdf
except ImportError:  # pragma: no cover
    pymupdf = None


ROOT = Path(__file__).resolve().parents[1]
RESULT_DIR = ROOT / "src" / "data" / "result"
OUT_DIR = ROOT / "src" / "data" / "processed"
OUT_FILE = OUT_DIR / "admissions.json"
REPORT_FILE = OUT_DIR / "extract_report.json"

NUM_RE = re.compile(r"^-?\d+(?:[.,]\d+)?%?$")
FILENAME_RE = re.compile(
    r"^(?P<univ>.+?)_(?P<year>\d{4})\s*(?P<kind>수시|정시|입시결과)?(?P<rest>.*)\.(?P<ext>xlsx|pdf)$",
    re.IGNORECASE,
)


@dataclass
class MajorRow:
    major: str
    track: str = ""
    campus: str = ""
    group: str = ""
    field: str = ""
    quotaInitial: float | None = None
    applicants: float | None = None
    competition: float | None = None
    enrolled: float | None = None
    waitlistNo: float | None = None
    cut50: float | None = None
    cut70: float | None = None
    cut80: float | None = None
    cut90: float | None = None
    cut100: float | None = None
    avg: float | None = None
    max: float | None = None
    min: float | None = None
    scoreKind: str = "unknown"


@dataclass
class UniversityDataset:
    id: str
    university: str
    year: int
    admissionType: str
    track: str
    scoreUnit: str
    scoreKind: str
    source: str
    sourceNote: str
    majors: list[MajorRow] = field(default_factory=list)


def clean_text(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).replace("\xa0", " ").replace("\n", " ").strip()
    return re.sub(r"\s+", " ", s)


def to_number(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (v != v):  # NaN
            return None
        return float(v)
    s = clean_text(v)
    if not s or s in {"-", "—", "–", "N/A", "na", "없음", "."}:
        return None
    s = s.replace(",", "").replace("%", "").replace(" ", "")
    # OCR/excel quirks like 5 89 -> 5.89 already handled in sheets usually
    if not re.fullmatch(r"-?\d+(?:\.\d+)?", s):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_filename(path: Path) -> dict[str, Any]:
    m = FILENAME_RE.match(path.name)
    if not m:
        return {
            "university": path.stem,
            "year": 0,
            "admissionType": "기타",
            "campusHint": "",
        }
    rest = m.group("rest") or ""
    kind = m.group("kind") or ""
    if kind == "입시결과":
        # e.g. 성신여대_2026 입시결과.pdf — unknown season
        if "수시" in rest:
            admission = "수시"
        elif "정시" in rest:
            admission = "정시"
        else:
            admission = "통합"
    elif kind in {"수시", "정시"}:
        admission = kind
    else:
        admission = "정시" if "정시" in path.name else ("수시" if "수시" in path.name else "기타")

    campus = ""
    cm = re.search(r"\(([^)]+)\)", path.stem)
    if cm:
        campus = cm.group(1)

    return {
        "university": m.group("univ").strip(),
        "year": int(m.group("year")),
        "admissionType": admission,
        "campusHint": campus,
    }


def score_kind_from_headers(headers: list[str]) -> str:
    joined = " ".join(headers)
    if "등급" in joined and "백분위" not in joined and "환산" not in joined:
        return "grade"
    if "백분위" in joined:
        return "percentile"
    if "환산" in joined or "컷" in joined or "cut" in joined.lower():
        return "converted"
    return "mixed"


def classify_header(cell: str) -> str | None:
    h = clean_text(cell).lower().replace(" ", "")
    if not h:
        return None
    mapping = [
        (("모집단위", "학과명", "모집학과", "학부", "전공명"), "major"),
        (("모집전형", "전형명", "전형구분", "세부전형", "전형"), "track"),
        (("캠퍼스",), "campus"),
        (("계열", "단과대학", "대학"), "field"),
        (("모집군", "군"), "group"),
        (("모집인원", "모집\n인원", "정원"), "quota"),
        (("지원인원", "지원자", "지원\n인원"), "applicants"),
        (("경쟁률",), "competition"),
        (("최종등록", "등록인원", "등록\n인원"), "enrolled"),
        (("예비", "충원"), "waitlist"),
        (("50%cut", "50%컷", "등급(50%)", "학생부등급(50%)"), "cut50"),
        (("70%cut", "70%컷", "등급(70%)", "학생부등급(70%)", "75%컷", "75%cut"), "cut70"),
        (("80%컷", "80%cut", "등록80%", "등록 80%"), "cut80"),
        (("90%컷", "90%cut", "등급(90%)"), "cut90"),
        (("100%컷", "100%cut"), "cut100"),
        (("평균", "평균백분위", "환산점수"), "avg"),
        (("최고", "최우수"), "max"),
        (("최저", "최하"), "min"),
        (("백분위",), "percentile_generic"),
        (("수능등급", "등급"), "grade_generic"),
    ]
    # exact-ish contains checks with priority order
    for keys, label in mapping:
        for k in keys:
            kk = k.lower().replace(" ", "").replace("\n", "")
            if kk in h:
                return label
    return None


def merge_header_rows(row_a: list[Any], row_b: list[Any] | None) -> list[str]:
    width = max(len(row_a), len(row_b or []))
    merged: list[str] = []
    for j in range(width):
        a = clean_text(row_a[j]) if j < len(row_a) else ""
        b = clean_text(row_b[j]) if row_b is not None and j < len(row_b) else ""
        if a and b and a != b:
            merged.append(f"{a} {b}")
        else:
            merged.append(a or b)
    return merged


def find_header_row(rows: list[list[Any]], max_scan: int = 25) -> tuple[int, dict[str, int], list[str]] | None:
    best = None
    for i, row in enumerate(rows[:max_scan]):
        candidates = [([clean_text(c) for c in row], i)]
        if i + 1 < len(rows):
            candidates.append((merge_header_rows(row, rows[i + 1]), i + 1))

        for headers, data_start in candidates:
            labels: dict[str, int] = {}
            for j, cell in enumerate(headers):
                kind = classify_header(cell)
                if kind and kind not in labels:
                    labels[kind] = j
            for j, cell in enumerate(headers):
                h = cell.replace(" ", "")
                if "최종등록" in h and "50%" in h:
                    labels["cut50"] = j
                if "최종등록" in h and ("70%" in h or "75%" in h):
                    labels["cut70"] = j
                if "최종등록자" in h and h.endswith("평균"):
                    labels.setdefault("avg", j)
            score = 0
            if "major" in labels:
                score += 3
            for k in ("competition", "cut70", "cut50", "avg", "quota", "applicants", "cut80"):
                if k in labels:
                    score += 1
            if score >= 3 and "major" in labels:
                rank = (score, -data_start)
                if best is None or rank > best[0]:
                    best = (rank, data_start, labels, headers)
    if not best:
        return None
    _, idx, labels, headers = best
    return idx, labels, headers


def extract_sheet_rows(ws) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def row_to_major(row: list[Any], colmap: dict[str, int], headers: list[str], defaults: dict[str, str]) -> MajorRow | None:
    def get(label: str) -> Any:
        idx = colmap.get(label)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    major = clean_text(get("major"))
    if not major or major in {"합계", "소계", "계", "모집단위", "학과", "구분"}:
        return None
    if re.fullmatch(r"\d+", major):
        return None

    # skip mostly empty numeric rows without a real major-looking string
    if len(major) < 2:
        return None

    track = clean_text(get("track")) or defaults.get("track", "")
    campus = clean_text(get("campus")) or defaults.get("campus", "")
    group = clean_text(get("group")) or defaults.get("group", "")
    field = clean_text(get("field")) or defaults.get("field", "")

    cut70 = to_number(get("cut70"))
    cut50 = to_number(get("cut50"))
    cut80 = to_number(get("cut80"))
    cut90 = to_number(get("cut90"))
    cut100 = to_number(get("cut100"))
    avg = to_number(get("avg"))
    mx = to_number(get("max"))
    mn = to_number(get("min"))

    # fallbacks from generic columns
    if cut70 is None and "percentile_generic" in colmap:
        # sometimes only one percentile column = 80% cut or avg
        pass
    if avg is None and "percentile_generic" in colmap:
        avg = to_number(get("percentile_generic"))
    if cut80 is None and "percentile_generic" in colmap and "cut80" not in colmap:
        # 호서대 style: 백분위 (80% 컷)
        hdr = headers[colmap["percentile_generic"]] if colmap["percentile_generic"] < len(headers) else ""
        if "80" in hdr or "컷" in hdr:
            cut80 = to_number(get("percentile_generic"))

    competition = to_number(get("competition"))
    quota = to_number(get("quota"))
    applicants = to_number(get("applicants"))
    enrolled = to_number(get("enrolled"))
    waitlist = to_number(get("waitlist"))

    # Require at least one score-like metric or competition
    metrics = [cut50, cut70, cut80, cut90, cut100, avg, mx, mn, competition]
    if all(m is None for m in metrics):
        return None

    kind = score_kind_from_headers(headers)
    if cut70 is not None or cut50 is not None or cut80 is not None:
        if any("등급" in h for h in headers):
            kind = "grade"
        elif any("백분위" in h for h in headers):
            kind = "percentile"
        else:
            kind = "converted"

    return MajorRow(
        major=major,
        track=track,
        campus=campus,
        group=group.replace("군", "") if group.endswith("군") and len(group) <= 3 else group,
        field=field,
        quotaInitial=quota,
        applicants=applicants,
        competition=competition,
        enrolled=enrolled,
        waitlistNo=waitlist,
        cut50=cut50,
        cut70=cut70 if cut70 is not None else cut80,
        cut80=cut80,
        cut90=cut90,
        cut100=cut100 if cut100 is not None else mn,
        avg=avg,
        max=mx,
        min=mn,
        scoreKind=kind,
    )


def extract_xlsx(path: Path, meta: dict[str, Any]) -> tuple[list[MajorRow], str]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    all_majors: list[MajorRow] = []
    notes: list[str] = []
    for sheet_name in wb.sheetnames:
        if any(k in sheet_name for k in ("목차", "안내", "표지")):
            continue
        ws = wb[sheet_name]
        rows = extract_sheet_rows(ws)
        found = find_header_row(rows)
        if not found:
            notes.append(f"{sheet_name}: header not found")
            continue
        header_idx, colmap, headers = found
        defaults = {
            "track": sheet_name if sheet_name not in {"Sheet1", "Sheet2", "정시", "수시"} else meta["admissionType"],
            "campus": meta.get("campusHint", ""),
        }
        # carry-forward major/track when merged cells leave blanks
        last_major = ""
        last_track = defaults["track"]
        last_campus = defaults["campus"]
        last_field = ""
        count = 0
        for row in rows[header_idx + 1 :]:
            # fill forward blank major from previous if competition exists
            cells = list(row)
            mi = colmap.get("major")
            if mi is not None and mi < len(cells):
                if clean_text(cells[mi]):
                    last_major = clean_text(cells[mi])
                elif last_major and any(to_number(c) is not None for c in cells):
                    cells[mi] = last_major
            ti = colmap.get("track")
            if ti is not None and ti < len(cells):
                if clean_text(cells[ti]):
                    last_track = clean_text(cells[ti])
                else:
                    cells[ti] = last_track
            ci = colmap.get("campus")
            if ci is not None and ci < len(cells):
                if clean_text(cells[ci]):
                    last_campus = clean_text(cells[ci])
                else:
                    cells[ci] = last_campus
            fi = colmap.get("field")
            if fi is not None and fi < len(cells):
                if clean_text(cells[fi]):
                    last_field = clean_text(cells[fi])
                else:
                    cells[fi] = last_field

            defaults_row = {
                "track": last_track or defaults["track"],
                "campus": last_campus or defaults["campus"],
                "field": last_field,
                "group": "",
            }
            major = row_to_major(cells, colmap, headers, defaults_row)
            if major:
                all_majors.append(major)
                count += 1
        notes.append(f"{sheet_name}: {count} rows")
    wb.close()
    return all_majors, "; ".join(notes)


def extract_pdf_text_tables(path: Path, meta: dict[str, Any]) -> tuple[list[MajorRow], str]:
    if pymupdf is None:
        return [], "pymupdf not installed"
    doc = pymupdf.open(path)
    majors: list[MajorRow] = []
    table_count = 0

    for page_index in range(len(doc)):
        page = doc[page_index]
        try:
            found = page.find_tables()
            tables = list(getattr(found, "tables", found) or [])
        except Exception:  # noqa: BLE001
            tables = []

        for table in tables:
            table_count += 1
            try:
                raw_rows = table.extract()
            except Exception:  # noqa: BLE001
                continue
            if not raw_rows or len(raw_rows) < 2:
                continue

            # Build merged headers from first 1-3 rows
            header_guess_rows = raw_rows[:3]
            best = None
            for depth in (1, 2, 3):
                if depth > len(header_guess_rows):
                    break
                headers = header_guess_rows[0]
                for extra in header_guess_rows[1:depth]:
                    headers = merge_header_rows(headers, extra)
                labels: dict[str, int] = {}
                for j, cell in enumerate(headers):
                    kind = classify_header(clean_text(cell))
                    if kind and kind not in labels:
                        labels[kind] = j
                # Common PDF layouts: 최저/평균/최고 as cuts
                for j, cell in enumerate(headers):
                    h = clean_text(cell)
                    if h == "최저" or h.endswith(" 최저"):
                        labels.setdefault("cut70", j)  # many schools publish 최저 as practical cut
                        labels.setdefault("min", j)
                    if h == "평균" or h.endswith(" 평균"):
                        labels.setdefault("avg", j)
                    if h == "최고" or h.endswith(" 최고"):
                        labels.setdefault("max", j)
                score = (3 if "major" in labels else 0) + sum(
                    1 for k in ("competition", "cut70", "avg", "quota", "applicants") if k in labels
                )
                if score >= 3 and "major" in labels:
                    best = (score, depth, labels, [clean_text(c) for c in headers])
            if not best:
                continue
            _, depth, colmap, headers = best
            defaults = {
                "track": meta["admissionType"],
                "campus": meta.get("campusHint", ""),
            }
            last_major = ""
            last_track = defaults["track"]
            last_field = ""
            last_campus = defaults["campus"]
            for row in raw_rows[depth:]:
                cells = list(row)
                mi = colmap.get("major")
                if mi is not None and mi < len(cells):
                    if clean_text(cells[mi]):
                        last_major = clean_text(cells[mi])
                    elif last_major:
                        cells[mi] = last_major
                ti = colmap.get("track")
                if ti is not None and ti < len(cells):
                    if clean_text(cells[ti]):
                        last_track = clean_text(cells[ti])
                    else:
                        cells[ti] = last_track
                fi = colmap.get("field")
                if fi is not None and fi < len(cells):
                    if clean_text(cells[fi]):
                        last_field = clean_text(cells[fi])
                    else:
                        cells[fi] = last_field
                ci = colmap.get("campus")
                if ci is not None and ci < len(cells):
                    if clean_text(cells[ci]):
                        last_campus = clean_text(cells[ci])
                    else:
                        cells[ci] = last_campus
                major = row_to_major(
                    cells,
                    colmap,
                    headers,
                    {"track": last_track, "campus": last_campus, "field": last_field, "group": ""},
                )
                if major:
                    majors.append(major)

    doc.close()
    # Deduplicate identical major+track+cut70
    uniq: list[MajorRow] = []
    seen: set[tuple] = set()
    for m in majors:
        key = (m.major, m.track, m.campus, m.cut70, m.competition)
        if key in seen:
            continue
        seen.add(key)
        uniq.append(m)

    note = f"pdf tables={table_count}, extracted={len(uniq)}"
    return uniq, note


def dominant_score_kind(majors: list[MajorRow]) -> str:
    counts: dict[str, int] = {}
    for m in majors:
        counts[m.scoreKind] = counts.get(m.scoreKind, 0) + 1
    if not counts:
        return "unknown"
    return max(counts.items(), key=lambda x: x[1])[0]


def load_existing_kyonggi() -> UniversityDataset | None:
    path = ROOT / "src" / "data" / "kyonggi-2026-jeongsi.json"
    if not path.exists():
        return None
    raw = json.loads(path.read_text(encoding="utf-8"))
    majors = [MajorRow(**{k: m.get(k) for k in MajorRow.__dataclass_fields__ if k in m}) for m in raw.get("majors", [])]
    for m in majors:
        if not m.scoreKind or m.scoreKind == "unknown":
            m.scoreKind = "percentile"
    return UniversityDataset(
        id="경기대-2026-정시",
        university="경기대",
        year=int(raw.get("year", 2026)),
        admissionType=raw.get("admissionType", "정시"),
        track=raw.get("track", "수능(일반학생전형)"),
        scoreUnit=raw.get("scoreUnit", "환산 백분위"),
        scoreKind="percentile",
        source=raw.get("source", path.name),
        sourceNote=raw.get("sourceNote", "기존 적재분"),
        majors=majors,
    )


def process_all() -> dict[str, Any]:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    datasets: list[UniversityDataset] = []
    report: list[dict[str, Any]] = []

    files = sorted(RESULT_DIR.glob("*"))
    files = [p for p in files if p.suffix.lower() in {".xlsx", ".pdf"}]
    # 수시 입결은 제외 (정시만 적재)
    files = [p for p in files if "수시" not in p.name]

    for path in files:
        meta = parse_filename(path)
        try:
            if path.suffix.lower() == ".xlsx":
                majors, note = extract_xlsx(path, meta)
            else:
                majors, note = extract_pdf_text_tables(path, meta)
            status = "ok" if majors else "empty"
        except Exception as exc:  # noqa: BLE001
            majors, note, status = [], f"error: {exc}", "error"

        ds_id = f"{meta['university']}-{meta['year']}-{meta['admissionType']}"
        if meta.get("campusHint"):
            ds_id += f"-{meta['campusHint']}"

        kind = dominant_score_kind(majors)
        unit = {
            "grade": "학생부/수능 등급",
            "percentile": "백분위",
            "converted": "대학 환산점수",
            "mixed": "혼합",
            "unknown": "미상",
        }.get(kind, "미상")

        ds = UniversityDataset(
            id=ds_id,
            university=meta["university"],
            year=meta["year"],
            admissionType=meta["admissionType"],
            track=meta["admissionType"],
            scoreUnit=unit,
            scoreKind=kind,
            source=path.name,
            sourceNote=note,
            majors=majors,
        )
        if majors:
            datasets.append(ds)
        report.append(
            {
                "file": path.name,
                "status": status,
                "majors": len(majors),
                "note": note,
                "admissionType": meta["admissionType"],
                "year": meta["year"],
                "university": meta["university"],
            }
        )

    kyonggi = load_existing_kyonggi()
    if kyonggi and not any(d.university.startswith("경기") for d in datasets):
        datasets.insert(0, kyonggi)
        report.append(
            {
                "file": kyonggi.source,
                "status": "ok",
                "majors": len(kyonggi.majors),
                "note": "merged existing kyonggi json",
                "admissionType": kyonggi.admissionType,
                "year": kyonggi.year,
                "university": kyonggi.university,
            }
        )

    payload = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceDir": str(RESULT_DIR),
        "universityCount": len(datasets),
        "majorCount": sum(len(d.majors) for d in datasets),
        "universities": [
            {
                **{k: v for k, v in asdict(d).items() if k != "majors"},
                "majors": [asdict(m) for m in d.majors],
            }
            for d in datasets
        ],
    }
    OUT_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    REPORT_FILE.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"payload": payload, "report": report}


def main() -> int:
    if not RESULT_DIR.exists():
        print(f"missing result dir: {RESULT_DIR}", file=sys.stderr)
        return 1
    result = process_all()
    payload, report = result["payload"], result["report"]
    ok = sum(1 for r in report if r["status"] == "ok")
    empty = sum(1 for r in report if r["status"] == "empty")
    err = sum(1 for r in report if r["status"] == "error")
    print(f"wrote {OUT_FILE}")
    print(f"universities={payload['universityCount']} majors={payload['majorCount']}")
    print(f"files ok={ok} empty={empty} error={err}")
    for r in report:
        print(f"- [{r['status']}] {r['file']}: {r['majors']} ({r['note'][:80]})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
